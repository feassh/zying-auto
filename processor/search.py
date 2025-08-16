import math
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import List, Optional, TYPE_CHECKING, Any

import pyperclip
import requests
import win32gui
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.styles.builtins import total
from openpyxl.utils import get_column_letter
from pywinauto import WindowSpecification
from pywinauto.application import Application
from rich.progress import Progress, BarColumn, TextColumn, TimeRemainingColumn

import config
import util

if TYPE_CHECKING:
    from process import AsyncWorker  # 仅类型检查时导入


def sleep(seconds: float):
    time.sleep(seconds)


class SearchProcessor:
    def __init__(self, worker: 'AsyncWorker'):
        self._worker = worker
        self.saved_kw_number = 0

    def log(self, content: str, color="black"):
        self._worker.log_signal.emit(f"{content}\n", color)

    def log_debug(self, content: str, color="white"):
        self._worker.log_debug_signal.emit(f"{content}\n\n", color)

    def progress(self, value: int):
        self._worker.progress_signal.emit(value)

    def update_current_page(self, value: int):
        self._worker.page_signal.emit(value)

    def update_saved_number(self, value: int):
        self.saved_kw_number += value
        self._worker.saved_number_signal.emit(self.saved_kw_number)

    def app_login(self) -> Application:
        # 先结束已打开的所有 ZYing.exe 进程
        util.system.kill_process_by_name("ZYing.exe")

        self.log("正在启动目标应用程序...")
        app = Application(backend="uia").start(config.get_config()['exePath'])

        # 等待登录窗口出现
        while True:
            if win32gui.FindWindow("WindowsForms10.Window.20008.app.0.34f5582_r3_ad1", None) != 0:
                break

        # 连接到窗口
        login_window = app.window(title="系统登录")

        # 等待窗口出现（重要！）
        login_window.wait("visible", timeout=300)

        self.log("正在登录...", "orange")
        input_username = login_window.child_window(auto_id="txtAcc")
        input_pwd = login_window.child_window(auto_id="txtPwd")

        if input_username.window_text() == "":
            input_username.set_focus()
            input_username.type_keys(config.get_config()['user'])

        input_pwd.set_focus()
        input_pwd.type_keys(config.get_config()['pwd'] + "{ENTER}", with_spaces=True)

        return app

    def navigate_to_target_page(self, auto_click=True) -> WindowSpecification:
        self.log("正在导航至目标页面...", "orange")

        app = Application(backend="win32").connect(title="分销系统", timeout=300)

        main_window = app.window(title="分销系统")
        main_window.wait("visible", timeout=300)

        self.log("正在初始化应用程序窗口...")
        main_window.set_focus()
        main_window.move_window(x=0, y=0, width=1110, height=700, repaint=True)
        sleep(1)

        if not auto_click:
            return main_window

        steps = [
            ("点击【产品】选项卡", 296, 40),
            ("点击【亚马逊选品】选项卡", 56, 678),
            ("点击【搜索词排名】选项卡", 71, 618),
            ("点击【日本】选项卡", 329, 77),
            ("点击【50000+】筛选条件", 747, 159),
            ("点击【排名上升+】筛选条件", 494, 265),
            ("点击【1万+】筛选条件", 655, 299),
        ]

        for desc, x, y in steps:
            self.log(desc)
            util.system.safe_click(x, y)
            util.app.check_load_finished()

        return main_window

    def get_total_kw_page_and_item(self, window: WindowSpecification) -> tuple[int, int]:
        total_item = 0
        total_page = 0

        match = re.search(r'共有\s*(\d+)\s*条记录', window.child_window(auto_id="lblTotal").window_text())
        if match:
            total_item = int(match.group(1))
            total_page = math.ceil(total_item / 60)

        return total_item, total_page

    def get_amazon_cookies(self) -> str:
        """从亚马逊获取最新的Cookie。如果失败，则使用硬编码的备用Cookie。"""
        try:
            self.log("正在尝试获取最新的亚马逊 Cookie...")

            # 先获取 session-id
            res_session_id = util.net.get(
                "https://www.amazon.co.jp/s?k=cat",
                headers={
                    "Content-Type": "text/html;charset=UTF-8",
                    "User-Agent": config.user_agent,
                },
                stream=True
            )

            # 获取 anti-csrftoken-a2z
            csrf_token = ''
            try:
                match = re.search(r'&quot;anti-csrftoken-a2z&quot;:&quot;(.*?)&quot;', res_session_id.text)
                if match:
                    source_csrf_token = match.group(1)

                    res_rendered_address_selections = util.net.get(
                        "https://www.amazon.co.jp/portal-migration/hz/glow/get-rendered-address-selections?deviceType=desktop&pageType=Search&storeContext=NoStoreName&actionSource=desktop-modal",
                        headers={
                            "Content-Type": "text/html;charset=UTF-8",
                            "Referer": "https://www.amazon.co.jp/s?k=cat",
                            "User-Agent": config.user_agent,
                            "Cookie": res_session_id.headers.get("set-cookie"),
                            "anti-csrftoken-a2z": source_csrf_token,
                        }
                    )

                    match = re.search(r'CSRF_TOKEN\s*:\s*"([^"]+)"', res_rendered_address_selections.text)
                    if match:
                        csrf_token = match.group(1)
            except:
                pass

            # 再获取 ubid-acbjp
            res_ubid = util.net.post(
                "https://www.amazon.co.jp/portal-migration/hz/glow/address-change?actionSource=glow",
                json_data={
                    "locationType": "LOCATION_INPUT",
                    "zipCode": "169-0074",
                    "deviceType": "web",
                    "storeContext": "hpc",
                    "pageType": "Detail",
                    "actionSource": "glow"
                },
                headers={
                    "Content-Type": "application/json",
                    "User-Agent": config.user_agent,
                    "Cookie": res_session_id.headers.get("set-cookie"),
                    # 这个参数是用来获取该接口返回的 json 数据的，如果不传此参数，可以正常获取包含地址的 Cookie 数据，但服务器会返回空
                    "anti-csrftoken-a2z": csrf_token,
                },
                stream=True
            )

            session_id = res_session_id.cookies.get("session-id", "")
            ubid = res_ubid.cookies.get("ubid-acbjp", "")

            if len(session_id) == 0 or len(ubid) == 0:
                raise Exception("Response is ok, but values are empty")
            else:
                try:
                    addr_data = res_ubid.json().get("address", {})
                    country = util.region.country_code_to_chinese(addr_data.get("countryCode", ""))
                    state = addr_data.get("state", "")
                    city = addr_data.get("city", "")
                    district = addr_data.get("district", "")
                    self.log(f"当前选择的配送地址：{country} {state} {city} {district}", "green")
                except:
                    pass

                return "ubid-acbjp=" + ubid + "; session-id=" + session_id
        except Exception as e:
            self.log(f"Cookie 获取失败: {e}", "red")
            self.log("将使用备用的 Cookie 数据。", "orange")

            # 从浏览器开发者工具中拿到的 Cookie 数据，有效期一年
            return "ubid-acbjp=355-5685452-2837352; session-id=357-7564356-4927846"

    def get_product_fulfiller_type(self, soup: BeautifulSoup) -> Optional[int]:
        delivery_1_el = soup.select_one('div#fulfillerInfoFeature_feature_div')
        delivery_2_el = soup.select_one('div#DELIVERY_JP')

        if delivery_1_el or delivery_2_el:
            if delivery_1_el:
                delivery_1_content = delivery_1_el.get_text()
                if "Amazon" in delivery_1_content:
                    # FBA
                    return 0

            if delivery_2_el:
                delivery_2_content = delivery_2_el.get_text()
                if "Amazon" in delivery_2_content:
                    # FBA
                    return 0

            # FBM
            return 1
        else:
            # 未知配送方式
            return None

    def process_product_detail(self, product: tuple, kw: str, cookies: str) -> Optional[tuple[tuple, tuple]]:
        asin, *_ = product

        try:
            res = util.net.get(
                f"https://www.amazon.co.jp/dp/{asin}?language=zh_CN",
                headers={
                    "User-Agent": config.user_agent,
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    "Accept-Encoding": "gzip, deflate, br, zstd",
                    "Cookie": cookies,
                }
            )

            soup = BeautifulSoup(res.text, "html.parser")
            fulfiller_type = self.get_product_fulfiller_type(soup)
            unknown_category = "未知分类"

            # 获取分类
            breadcrumbs_element = soup.select_one("div#wayfinding-breadcrumbs_feature_div")
            if not breadcrumbs_element:
                self.log_debug(f"搜索词 {kw} 的商品：{asin} 分类数据未获取到", "orange")
                return (unknown_category, "", fulfiller_type), product

            category_elements = breadcrumbs_element.select("a.a-link-normal.a-color-tertiary")
            categories = []
            for category_el in category_elements:
                cate = category_el.get_text().strip()
                if len(cate) == 0:
                    continue

                categories.append(cate)

            if len(categories) == 0:
                self.log_debug(f"搜索词 {kw} 的商品：{asin} 分类数据为空")
                return (unknown_category, "", fulfiller_type), product

            return (categories[0], '›'.join(map(str, categories[1:])), fulfiller_type), product
        except Exception as e:
            self.log_debug(f"搜索词 {kw} 的商品：{asin} 详情页获取失败：{e}", "red")
            return None

    def process_keyword(self, kw: str, cookies: str) -> Optional[tuple]:
        """为单个关键词从亚马逊获取并解析数据。"""
        try:
            res = util.net.get(
                f"https://www.amazon.co.jp/s?k={kw.strip()}&language=zh_CN",
                headers={
                    "User-Agent": config.user_agent,
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    "Accept-Encoding": "gzip, deflate, br, zstd",
                    "Cookie": cookies,
                }
            )

            soup = BeautifulSoup(res.text, "html.parser")

            # 检查是否被亚马逊风控
            if "ご迷惑をおかけしています" in soup.title.string:
                self.log("** 遇到亚马逊风控，操作终止 **", "red")
                return None

            if "Amazon.co.jp" not in soup.title.string:
                self.log("注意，获取的网页结果可能存在问题", "orange")

            # elements = soup.select('div[cel_widget_id*="MAIN-SEARCH_RESULTS-"]')
            elements = soup.select('div[data-component-type="s-search-result"]')
            valuable_els = []
            today = date.today()

            for el in elements:
                delivery_message = el.select_one("div.udm-primary-delivery-message span.a-text-bold")
                if not delivery_message:
                    continue

                match = re.search(r'(\d{1,2})月(\d{1,2})日', delivery_message.get_text())
                if not match:
                    continue

                month, day = int(match.group(1)), int(match.group(2))

                try:
                    target_date = date(today.year, month, day)
                    offset = (target_date - today).days
                    if config.get_config()['minDateInterval'] <= offset <= config.get_config()['maxDateInterval']:
                        valuable_els.append(el)
                except ValueError:
                    continue  # 忽略无效日期，如2月30日

            if len(valuable_els) < config.get_config()['matchCount']:
                return None

            kw_img = None
            products = []

            for ve in valuable_els:
                img = None
                title = ''
                price_symbol = None
                price = None
                buy_number = None

                product_asin = ve['data-asin'] if 'data-asin' in ve.attrs else None
                if product_asin and len(product_asin.strip()) > 0:
                    img_el = ve.find('img', class_='s-image')
                    img_result = img_el['src'] if img_el and 'src' in img_el.attrs else None
                    if img_result is not None and img_result.strip().startswith('http'):
                        img = img_result

                    title_el = ve.select_one('div[data-cy="title-recipe"]')
                    if title_el:
                        h2_el = title_el.find('h2')
                        if h2_el:
                            title = h2_el.get_text().strip()

                    price_el = ve.select_one('a[aria-describedby="price-link"]')
                    if price_el:
                        p_symbol_el = price_el.find('span', class_='a-price-symbol')
                        p_symbol = p_symbol_el.get_text().strip() if p_symbol_el else ''
                        if len(p_symbol) > 0:
                            price_symbol = p_symbol

                        p_whole_el = price_el.find('span', class_='a-price-whole')
                        if p_whole_el:
                            p_whole = p_whole_el.get_text().strip()
                            p_decimal_el = p_whole_el.find('span', class_='a-price-decimal')
                            p_decimal = p_decimal_el.text.strip() if p_decimal_el else ''
                            try:
                                price = int(p_whole.replace(',', '').replace(p_decimal, ''))
                            except Exception:
                                pass

                    buy_number_el = ve.select_one('div[data-cy="reviews-block"]')
                    if buy_number_el:
                        b_number_el = buy_number_el.select_one('span.a-size-base.a-color-secondary')
                        b_number_full = b_number_el.get_text().strip() if b_number_el else ''

                        # 匹配模式：
                        # (\d+)            —— 数字部分（整数）
                        # \s*              —— 允许数字和单位之间有空格
                        # (万|百万|千万|亿)? —— 单位，可选
                        match = re.search(r'过去一个月有(\d+)\s*(万|百万|千万|亿)?\+?', b_number_full)
                        if match:
                            try:
                                num_str, unit = match.groups()
                                num = int(num_str)

                                # 单位换算表
                                unit_map = {
                                    None: 1,
                                    '万': 10_000,
                                    '百万': 1_000_000,
                                    '千万': 10_000_000,
                                    '亿': 100_000_000
                                }

                                factor = unit_map.get(unit, 1)
                                buy_number = int(num * factor)
                            except Exception:
                                pass

                    products.append((product_asin.strip(), img, title, price_symbol, price, buy_number))
                else:
                    self.log_debug(f"搜索词 {kw} 未找到对应商品的 asin 数据", "orange")

                # 如果搜索词需要的图片获取到了，就不用执行下面的代码了
                if kw_img is not None:
                    continue

                img_el = ve.find('img', class_='s-image')
                img_result = img_el['src'] if img_el and 'src' in img_el.attrs else None
                if img_result is not None and img_result.strip().startswith('http'):
                    kw_img = img_result

            concurrency = config.get_config()['concurrency']
            self.log_debug(f"获取到 {len(products)} 个符合筛选条件的商品，正在使用 {concurrency} 个并发线程进行处理...")

            saved_products = []

            with ThreadPoolExecutor(max_workers=concurrency) as executor:
                future_to_product = {executor.submit(self.process_product_detail, product, kw, cookies): product for
                                     product in products}
                for future in as_completed(future_to_product):
                    if self._worker.is_stopping():
                        self.log_debug("收到停止信号，正在终止任务...", "orange")
                        executor.shutdown(wait=True, cancel_futures=True)
                        break

                    result = future.result()
                    if result:
                        saved_products.append(result)

                        (cate_main, cate_sub, _), _ = result
                        self.log_debug(f"搜索词 {kw} 对应的的商品主分类：{cate_main}，次分类：{cate_sub}", "cyan")

            self.log_debug(f"搜索词 {kw.strip()} 获取到了 {len(saved_products)} 个商品和对应的商品数据",
                           "green" if len(saved_products) > 0 else "orange")

            filter_criteria = f"{config.get_config()['minDateInterval']}-{config.get_config()['maxDateInterval']}-{config.get_config()['matchCount']}"

            self.log_debug(f"符合条件搜索词：{kw.strip()}\n图片：{kw_img}", color="green")
            return kw.strip(), kw_img, filter_criteria, saved_products
        except requests.exceptions.RequestException as e:
            # 重试机制会处理此问题，但如果所有重试都失败，最好记录下来。
            self.log_debug(f"关键词 '{kw}' 请求失败: {e}", "red")
            return None
        except Exception as e:
            self.log_debug(f"处理关键词 '{kw}' 时出错: {e}", "red")
            return None

    def process_page_concurrently(
            self,
            window: Optional[WindowSpecification],
            page_rect: Optional[Any],
            cur_page: int,
            total_pages: int,
            cookies: str,
    ) -> List[tuple]:
        """处理单个页面的UI交互，获取关键词列表，并并发处理它们。"""
        self.log(f"\n--- 正在获取第 {cur_page + 1}/{total_pages} 页搜索词列表 ---")
        self.update_current_page(cur_page + 1)

        kw_list = None

        if config.is_zying_data_source():
            window.set_focus()
            sleep(1)

            # 点击一下空白区域，让页码按钮从 hover 状态退出，防止接下来 OpenCV 轮廓识别错误
            util.system.safe_click(x=541, y=647)
            sleep(1)

            # 导航到正确的页面
            if cur_page == 0:
                util.system.safe_click(x=155, y=600)
            else:
                while True:
                    point = util.cv.get_next_page_point(page_rect)
                    if point:
                        util.system.safe_click(x=point[0], y=point[1])
                        break
                    else:
                        sleep(1)

            util.app.check_load_finished()

            # 复制关键词到剪贴板
            pyperclip.copy("")
            util.system.safe_right_click(x=233, y=464)
            sleep(0.5)
            util.system.safe_click(x=287, y=504)
            sleep(0.5)

            kw_list = [kw for kw in pyperclip.paste().splitlines() if kw]
        else:
            list_data, _, _ = self.get_data_by_amz123(cur_page + 1)
            kw_list = [kw['word'] for kw in list_data if kw.get('word')]

        if not kw_list or len(kw_list) == 0:
            self.log("此页未找到关键词。", "orange")
            return []

        concurrency = config.get_config()['concurrency']
        self.log(f"获取到 {len(kw_list)} 个搜索词，正在使用 {concurrency} 个并发线程进行处理...")

        saved_kw = []

        # 使用 rich.progress 创建一个进度条
        with Progress(
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                "[progress.percentage]{task.percentage:>3.0f}%",
                TimeRemainingColumn(),
        ) as progress:
            task = progress.add_task("[cyan]正在筛选...", total=len(kw_list))

            # 使用线程池执行并发任务
            with ThreadPoolExecutor(max_workers=concurrency) as executor:
                # 提交所有任务
                future_to_kw = {executor.submit(self.process_keyword, kw, cookies): kw for kw in kw_list}

                # 获取已完成任务的结果
                for future in as_completed(future_to_kw):
                    if self._worker.is_stopping():
                        self.log("收到停止信号，正在终止任务...", "orange")
                        executor.shutdown(wait=True, cancel_futures=True)
                        break

                    result = future.result()
                    if result:
                        saved_kw.append(result)

                    progress.update(task, advance=1)  # 更新进度条
                    current_percentage = int(progress.tasks[task].percentage)
                    self.progress(current_percentage)

        return saved_kw

    def save_results(self, saved_kw: List[tuple], cur_page: int):
        """将处理后的关键词数据保存到Excel文件。"""
        if not saved_kw:
            self.log("当前页筛选后的搜索词数量为 0，跳过生成Excel文件。", "orange")
            return

        self.log(f"正在为 {len(saved_kw)} 个搜索词生成Excel文件...", "orange")

        # 目标目录
        output_dir = Path(config.get_config().get('excelPath') or Path(util.system.get_exe_dir()) / "excel")
        output_dir.mkdir(parents=True, exist_ok=True)

        # 文件名（按小时分）
        filename = datetime.now().strftime("%m-%d_%H") + ".xlsx"
        file_path = output_dir / filename

        is_new_file = not file_path.exists()

        # 打开或新建
        if file_path.exists():
            wb = load_workbook(file_path)
            # if f"第{cur_page + 1}页" in wb.sheetnames:
            if "亚马逊关键词" in wb.sheetnames:
                # ws = wb[f"第{cur_page + 1}页"]
                ws = wb["亚马逊关键词"]
            else:
                ws = wb.create_sheet("亚马逊关键词")
                # ws = wb.create_sheet(f"第{cur_page + 1}页")
                # ws['A1'] = "亚马逊 [日本] 区域关键词"
                # ws['A1'].font = Font(bold=True, size=20)
                # headers = ["关键词", "图片"]
                # for col, header in enumerate(headers, 1):
                #     ws.cell(row=2, column=col, value=header).font = Font(bold=True)
        else:
            wb = Workbook()
            ws = wb.active
            # ws.title = f"第{cur_page + 1}页"
            ws.title = "亚马逊关键词"
            ws['A1'] = "亚马逊 [日本] 区域关键词"
            ws['A1'].font = Font(bold=True, size=20)
            headers = ["关键词", "图片"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=2, column=col, value=header).font = Font(bold=True)

        # 找到追加的起始行
        start_row = ws.max_row + 1

        # 页码行
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        cell = ws.cell(row=start_row, column=1)
        cell.value = f"==== 第 {cur_page + 1} 页 ===="
        cell.font = Font(bold=True, size=20)

        # 数据起始行
        start_row += 1

        for i, (text, img_url, *_) in enumerate(saved_kw, start=start_row):
            cell = ws.cell(row=i, column=1, value=text)
            cell.hyperlink = f"https://www.amazon.co.jp/s?k={text}"
            cell.style = "Hyperlink"

            if img_url:
                try:
                    resp = util.net.get(img_url)
                    img = XLImage(BytesIO(resp.content))
                    img.width, img.height = 80, 80
                    ws.add_image(img, f"B{i}")
                    ws.row_dimensions[i].height = 65
                except Exception as e:
                    self.log(f"搜索词 {text} 的图片下载失败，已忽略: {img_url}\n{e}", "red")

        # 如果是新文件才设置列宽（避免覆盖旧文件调整过的宽度）
        if is_new_file:
            max_len = max((len(text) for text, *_ in saved_kw), default=0)
            ws.column_dimensions[get_column_letter(1)].width = max_len + 5
            ws.column_dimensions[get_column_letter(2)].width = 15

        try:
            wb.save(output_dir / filename)
            self.log(f"已将数据存储为 Excel 文件：{filename}", "green")
        except Exception as e:
            self.log(f"保存 Excel 文件失败: {e}", "red")

        # 数据上传到服务器
        self.log("正在将数据上传到服务器端...", "orange")
        ex = util.net.save_kw_to_server(saved_kw)
        if ex is None:
            self.log("** 数据已同步上传到服务器端 **", "green")
        else:
            self.log(f"** 数据上传服务器失败：{ex} **", "red")

        self.update_saved_number(len(saved_kw))

    def get_data_by_zying(self) -> tuple[list, tuple[int, int], Optional[WindowSpecification]]:
        # 根据配置决定是启动新应用还是连接现有应用
        config_current_page = config.get_config()['currentPage']
        if config_current_page <= 0:
            self.app_login()

        main_window = self.navigate_to_target_page(config_current_page <= 0)

        return [], self.get_total_kw_page_and_item(main_window), main_window

    def get_data_by_amz123(self, page=1) -> tuple[list, tuple[int, int], Optional[WindowSpecification]]:
        data, e = util.net.get_amz123_kw_list(page)
        if not data:
            self.log(f"从 【amz123】 获取数据失败：{e}", "red")
            return [], (0, 0), None

        kw_list, total = data

        return kw_list, (total, math.ceil(total / 200)), None

    def start_work(self):
        try:
            if config.is_zying_data_source():
                self.log("正在从 【智赢跨境】 获取搜索词数据...", "orange")
                data = self.get_data_by_zying()
            else:
                self.log("正在从 【amz123】 获取搜索词数据...", "orange")
                data = self.get_data_by_amz123()

            _, (total_item, total_page), main_window = data

            if total_item <= 0:
                self.log(f"\n搜索词分页数据获取失败！程序已终止运行。", "red")
                return
            else:
                self.log(f"共获取到 {total_item} 条搜索词数据，一共 {total_page} 页", "green")

            page_rect = main_window.child_window(auto_id="PTurn").rectangle() if config.is_zying_data_source() else None
            amazon_cookies = self.get_amazon_cookies()

            config_current_page = config.get_config()['currentPage']
            if config_current_page <= 0: config_current_page = 1

            for cur_page in range(config_current_page - 1, total_page):
                if self._worker.is_stopping():
                    self.log("收到停止信号，正在终止任务...", "orange")
                    break

                saved_keywords = self.process_page_concurrently(main_window, page_rect, cur_page, total_page,
                                                                amazon_cookies)

                if self._worker.is_stopping():
                    self.log("收到停止信号，任务在保存结果前被终止。", "orange")
                    break

                self.save_results(saved_keywords, cur_page)

            self.log("\n所有任务均已执行完毕！", "green")
        except Exception as e:
            self.log(f"\n程序出错，已终止运行！错误：{e}", "red")
