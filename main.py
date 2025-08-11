import ctypes
import math
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import List, Optional

import pyautogui
import pyperclip
import requests
import win32gui
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pywinauto import WindowSpecification
from pywinauto.application import Application
from rich.progress import Progress, BarColumn, TextColumn, TimeRemainingColumn

import config
import util


def setup_console_window():
    hwnd = ctypes.windll.kernel32.GetConsoleWindow()
    if hwnd:
        screen_width, screen_height = pyautogui.size()
        # SWP_NOZORDER=0x4 保持层次
        ctypes.windll.user32.SetWindowPos(hwnd, 0, 0, 700, 1110, screen_height - 700, 0x4)
    else:
        util.system.print_inline("找不到自身控制台窗口句柄，请手动将本软件窗口移至没有遮挡目标软件的地方", color="red")


def app_login() -> Application:
    # 先结束已打开的所有 ZYing.exe 进程
    util.system.kill_process_by_name("ZYing.exe")

    util.system.print_inline("正在启动目标应用程序...")
    app = Application(backend="uia").start(config.get_config()['exePath'])

    # 等待登录窗口出现
    while True:
        if win32gui.FindWindow("WindowsForms10.Window.20008.app.0.34f5582_r3_ad1", None) != 0:
            break

    # 连接到窗口
    login_window = app.window(title="系统登录")

    # 等待窗口出现（重要！）
    login_window.wait("visible", timeout=300)

    util.system.print_inline("正在登录...", color="yellow")
    input_username = login_window.child_window(auto_id="txtAcc")
    input_pwd = login_window.child_window(auto_id="txtPwd")

    if input_username.window_text() == "":
        input_username.set_focus()
        input_username.type_keys(config.get_config()['user'])

    input_pwd.set_focus()
    input_pwd.type_keys(config.get_config()['pwd'] + "{ENTER}", with_spaces=True)

    return app


def navigate_to_target_page(auto_click=True) -> WindowSpecification:
    util.system.print_inline("正在导航至目标页面...", color="blue")

    app = Application(backend="win32").connect(title="分销系统", timeout=300)

    main_window = app.window(title="分销系统")
    main_window.wait("visible", timeout=300)

    util.system.print_inline("正在初始化应用程序窗口...", color="green")
    main_window.set_focus()
    main_window.move_window(x=0, y=0, width=1110, height=700, repaint=True)
    time.sleep(1)

    if not auto_click:
        return main_window

    steps = [
        ("点击【产品】选项卡", 296, 40),
        ("点击【亚马逊选品】选项卡", 56, 678),
        ("点击【搜索词排名】选项卡", 71, 618),
        ("点击【日本】选项卡", 329, 77),
        ("点击【50000+】筛选条件", 747, 159),
        ("点击【排名上升+】筛选条件", 494, 265),
        ("点击【1万+】筛选条件", 655, 299),
    ]

    for desc, x, y in steps:
        util.system.print_inline(desc)
        util.system.safe_click(x, y)
        util.app.check_load_finished()

    return main_window


def get_total_kw_page_and_item(window: WindowSpecification) -> tuple[int, int]:
    total_item = 0
    total_page = 0

    match = re.search(r'共有\s*(\d+)\s*条记录', window.child_window(auto_id="lblTotal").window_text())
    if match:
        total_item = int(match.group(1))
        total_page = math.ceil(total_item / 60)

    return total_item, total_page


def get_amazon_cookies() -> str:
    """从亚马逊获取最新的Cookie。如果失败，则使用硬编码的备用Cookie。"""
    try:
        util.system.print_inline("正在尝试获取最新的亚马逊 Cookie...")

        # 先获取 session-id
        res_session_id = util.net.get(
            "https://www.amazon.co.jp/s?k=cat",
            headers={
                "Content-Type": "text/html;charset=UTF-8",
                "User-Agent": config.user_agent,
            },
            stream=True
        )

        # 再获取 ubid-acbjp
        res_ubid = util.net.post(
            "https://www.amazon.co.jp/portal-migration/hz/glow/address-change?actionSource=glow",
            json_data={
                "locationType": "LOCATION_INPUT",
                "zipCode": "169-0074",
                "deviceType": "web",
                "storeContext": "hpc",
                "pageType": "Detail",
                "actionSource": "glow"
            },
            headers={
                "Content-Type": "application/json",
                "User-Agent": config.user_agent,
                "Cookie": res_session_id.headers.get("set-cookie"),
            },
            stream=True
        )

        session_id = res_session_id.cookies.get("session-id", "")
        ubid = res_ubid.cookies.get("ubid-acbjp", "")
        if len(session_id) == 0 or len(ubid) == 0:
            raise Exception("Response is ok, but values are empty")
        else:
            return "ubid-acbjp=" + ubid + "; session-id=" + session_id
    except Exception as e:
        util.system.print_inline(f"Cookie 获取失败: {e}", color="red")
        util.system.print_inline("将使用备用的 Cookie 数据。", color="yellow")

        # 从浏览器开发者工具中拿到的 Cookie 数据，有效期一年
        return "ubid-acbjp=355-5685452-2837352; session-id=357-7564356-4927846"


def process_keyword(kw: str, cookies: str) -> Optional[tuple]:
    """为单个关键词从亚马逊获取并解析数据。"""
    try:
        res = util.net.get(
            f"https://www.amazon.co.jp/s?k={kw.strip()}&language=zh_CN",
            headers={
                "User-Agent": config.user_agent,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                "Accept-Encoding": "gzip, deflate, br, zstd",
                "Cookie": cookies,
            }
        )

        soup = BeautifulSoup(res.text, "html.parser")

        # 检查是否被亚马逊风控
        if "ご迷惑をおかけしています" in soup.title.string:
            util.system.print_inline("遇到亚马逊风控，操作终止", color="red")
            return None

        if "Amazon.co.jp" not in soup.title.string:
            util.system.print_inline("注意，获取的网页结果可能存在问题", color="yellow")

        delivery_messages = soup.select("div.udm-primary-delivery-message span.a-text-bold")

        count = 0
        today = date.today()
        for el in delivery_messages:
            match = re.search(r'(\d{1,2})月(\d{1,2})日', el.get_text())
            if not match:
                continue

            month, day = int(match.group(1)), int(match.group(2))

            try:
                target_date = date(today.year, month, day)
                if (target_date - today).days >= config.get_config()['minDateInterval']:
                    count += 1
            except ValueError:
                continue  # 忽略无效日期，如2月30日

        if count < config.get_config()['matchCount']:
            return None

        img = None
        price_symbol = None
        price = None
        buy_number = None

        elements = soup.select('div[cel_widget_id*="MAIN-SEARCH_RESULTS-"]')
        for element in elements:
            if img is not None and price_symbol is not None and price is not None and buy_number is not None:
                break

            if element.select('div.udm-primary-delivery-message span.a-text-bold') is None:
                continue

            if img is None:
                img_el = element.find('img', class_='s-image')
                img_result = img_el['src'] if img_el and 'src' in img_el.attrs else None
                if img_result is not None and img_result.strip().startswith('http'):
                    img = img_result

            if price_symbol is None or price is None:
                price_el = element.select_one('a[aria-describedby="price-link"]')
                if price_el:
                    if price_symbol is None:
                        p_symbol_el = price_el.find('span', class_='a-price-symbol')
                        p_symbol = p_symbol_el.get_text().strip() if p_symbol_el else ''
                        if len(p_symbol) > 0:
                            price_symbol = p_symbol

                    if price is None:
                        p_whole_el = price_el.find('span', class_='a-price-whole')
                        if p_whole_el:
                            p_whole = p_whole_el.get_text().strip()
                            p_decimal_el = p_whole_el.find('span', class_='a-price-decimal')
                            p_decimal = p_decimal_el.text.strip() if p_decimal_el else ''
                            try:
                                price = int(p_whole.replace(',', '').replace(p_decimal, ''))
                            except Exception:
                                pass

            if buy_number is None:
                buy_number_el = element.select_one('div[data-cy="reviews-block"]')
                if buy_number_el:
                    b_number_el = buy_number_el.select_one('span.a-size-base.a-color-secondary')
                    b_number_full = b_number_el.get_text().strip() if b_number_el else ''

                    # 匹配模式：
                    # (\d+)            —— 数字部分（整数）
                    # \s*              —— 允许数字和单位之间有空格
                    # (万|百万|千万|亿)? —— 单位，可选
                    match = re.search(r'过去一个月有(\d+)\s*(万|百万|千万|亿)?\+?', b_number_full)
                    if match:
                        try:
                            num_str, unit = match.groups()
                            num = int(num_str)

                            # 单位换算表
                            unit_map = {
                                None: 1,
                                '万': 10_000,
                                '百万': 1_000_000,
                                '千万': 10_000_000,
                                '亿': 100_000_000
                            }

                            factor = unit_map.get(unit, 1)
                            buy_number = int(num * factor)
                        except Exception:
                            pass

        return kw.strip(), img, price_symbol, price, buy_number
    except requests.exceptions.RequestException as e:
        # 重试机制会处理此问题，但如果所有重试都失败，最好记录下来。
        util.system.print_inline(f"关键词 '{kw}' 请求失败: {e}", color="red")
        return None
    except Exception as e:
        util.system.print_inline(f"处理关键词 '{kw}' 时出错: {e}", color="red")
        return None


def process_page_concurrently(
        window,
        page_rect,
        cur_page: int,
        total_pages: int,
        cookies: str
) -> List[tuple]:
    """处理单个页面的UI交互，获取关键词列表，并并发处理它们。"""
    util.system.print_inline(f"\n--- 正在获取第 {cur_page + 1}/{total_pages} 页搜索词列表 ---", style="bold")

    window.set_focus()
    time.sleep(1)

    # 点击一下空白区域，让页码按钮从 hover 状态退出，防止接下来 OpenCV 轮廓识别错误
    util.system.safe_click(x=541, y=647)
    time.sleep(1)

    # 导航到正确的页面
    if cur_page == 0:
        util.system.safe_click(x=155, y=600)
    else:
        while True:
            point = util.cv.get_next_page_point(page_rect)
            if point:
                util.system.safe_click(x=point[0], y=point[1])
                break
            else:
                time.sleep(1)

    util.app.check_load_finished()

    # 复制关键词到剪贴板
    pyperclip.copy("")
    util.system.safe_right_click(x=233, y=464)
    time.sleep(0.5)
    util.system.safe_click(x=287, y=504)
    time.sleep(0.5)

    kw_list = [kw for kw in pyperclip.paste().splitlines() if kw]

    if not kw_list:
        util.system.print_inline("此页未找到关键词。", color="yellow")
        return []

    concurrency = config.get_config()['concurrency']
    util.system.print_inline(f"获取到 {len(kw_list)} 个搜索词，正在使用 {concurrency} 个并发线程进行处理...")

    saved_kw = []

    # 使用 rich.progress 创建一个进度条
    with Progress(
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            "[progress.percentage]{task.percentage:>3.0f}%",
            TimeRemainingColumn(),
    ) as progress:
        task = progress.add_task("[cyan]正在筛选...", total=len(kw_list))

        # 使用线程池执行并发任务
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            # 提交所有任务
            future_to_kw = {executor.submit(process_keyword, kw, cookies): kw for kw in kw_list}

            # 获取已完成任务的结果
            for future in as_completed(future_to_kw):
                result = future.result()
                if result:
                    saved_kw.append(result)
                progress.update(task, advance=1)  # 更新进度条

    return saved_kw


def save_results(saved_kw: List[tuple], cur_page: int):
    """将处理后的关键词数据保存到Excel文件。"""
    if not saved_kw:
        util.system.print_inline("当前页筛选后的搜索词数量为 0，跳过生成Excel文件。", color="yellow")
        return

    util.system.print_inline(f"正在为 {len(saved_kw)} 个关键词生成Excel文件...", color="yellow")

    wb = Workbook()
    ws = wb.active

    ws.title = "亚马逊关键词"
    ws['A1'] = "亚马逊 [日本] 区域关键词"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["关键词", "图片"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header).font = Font(bold=True)

    for i, (text, img_url, *_) in enumerate(saved_kw, start=3):
        cell = ws.cell(row=i, column=1, value=text)
        cell.hyperlink = f"https://www.amazon.co.jp/s?k={text}"
        cell.style = "Hyperlink"

        if img_url:
            try:
                resp = util.net.get(img_url)
                img = XLImage(BytesIO(resp.content))
                img.width, img.height = 80, 80
                ws.add_image(img, f"B{i}")
                ws.row_dimensions[i].height = 65
            except Exception as e:
                util.system.print_inline(f"搜索词 {text} 的图片下载失败，已忽略: {img_url}\n{e}", color="red")

    max_len = max(len(text) for text, *_ in saved_kw)
    ws.column_dimensions[get_column_letter(1)].width = max_len + 5
    ws.column_dimensions[get_column_letter(2)].width = 15

    try:
        output_dir = Path(config.get_config()['excelPath'] or Path(util.system.get_exe_dir()) / "excel")
        output_dir.mkdir(parents=True, exist_ok=True)
        filename = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + f"-第{cur_page + 1}页.xlsx"
        wb.save(output_dir / filename)

        util.system.print_inline(f"已将数据存储为 Excel 文件：{filename}", color="green")
    except Exception as e:
        util.system.print_inline(f"保存 Excel 文件失败: {e}", color="red")

    # 数据上传到服务器
    ex = util.net.save_kw_to_server(saved_kw)
    if ex is None:
        util.system.print_inline("** 数据已同步上传到服务器端 **", color="green")
    else:
        util.system.print_inline(f"** 数据上传服务器失败（不影响流程，可忽略）：{ex} **", color="yellow")


def main():
    config.DEBUG = config.get_config()["debug"]

    setup_console_window()

    # 根据配置决定是启动新应用还是连接现有应用
    config_current_page = config.get_config()['currentPage']
    if config_current_page <= 0:
        app_login()

    main_window = navigate_to_target_page(config_current_page <= 0)

    total_item, total_page = get_total_kw_page_and_item(main_window)

    if total_item <= 0:
        util.system.print_inline("搜索词分页数据获取失败！程序已终止运行。", color="red")
        util.app.press_any_key_exit()
    else:
        util.system.print_inline(f"共获取到 {total_item} 条搜索词数据，一共 {total_page} 页", color="green")

    amazon_cookies = get_amazon_cookies()
    page_rect = main_window.child_window(auto_id="PTurn").rectangle()
    if config_current_page <= 0: config_current_page = 1

    for cur_page in range(config_current_page - 1, total_page):
        saved_keywords = process_page_concurrently(main_window, page_rect, cur_page, total_page, amazon_cookies)
        save_results(saved_keywords, cur_page)

    util.system.print_inline("\n所有任务均已执行完毕！", color="green", style="bold")
    util.app.press_any_key_exit()


if __name__ == '__main__':
    util.app.ensure_start_by_self()

    try:
        main()
    except Exception as e:
        # console.print_exception(show_locals=True)
        print(e)
        util.system.print_inline("程序发生未预料的严重错误！已结束运行。", color="red")
        util.app.press_any_key_exit()
