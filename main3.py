import ctypes
import math
import os
import re
import sys
import time
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pyautogui
import pyperclip
import requests
import win32gui
from bs4 import BeautifulSoup
from pywinauto.application import Application
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

import util

# DEBUG = False
#
#
# def main():
    # # 获取自身控制台窗口句柄
    # hwnd = ctypes.windll.kernel32.GetConsoleWindow()
    # if hwnd:
    #     screen_width, screen_height = pyautogui.size()
    #     # SWP_NOZORDER=0x4 保持层次
    #     ctypes.windll.user32.SetWindowPos(hwnd, 0, 0, 700, 1110, screen_height - 700, 0x4)
    # else:
    #     util.system.print_inline("找不到自身控制台窗口句柄，请手动将本软件窗口移至没有遮挡目标软件的地方", color="red")
    #
    # ################## 读取配置文件 ##################
    # config, ok = util.app.load_config()
    # if not ok:
    #     util.system.print_inline("配置文件读取失败！" + str(config), color="red")
    #     util.app.press_any_key_exit()
    #     sys.exit()

    # DEBUG = config["debug"]

    # if config['currentPage'] <= 0:
        # ################## 启动应用程序 ##################
        # # 先结束已打开的所有 ZYing.exe 进程
        # util.system.kill_process_by_name("ZYing.exe")
        #
        # util.system.print_inline("正在启动目标应用程序...")
        # app = Application(backend="uia").start(config['exePath'])
        #
        # # 等待登录窗口出现
        # while True:
        #     if win32gui.FindWindow("WindowsForms10.Window.20008.app.0.34f5582_r3_ad1", None) != 0:
        #         break
        #
        # # 连接到窗口
        # login_window = app.window(title="系统登录")
        #
        # # 等待窗口出现（重要！）
        # login_window.wait("visible", timeout=300)
        #
        # util.system.print_inline("正在登录...", color="yellow")
        # input_username = login_window.child_window(auto_id="txtAcc")
        # input_pwd = login_window.child_window(auto_id="txtPwd")
        # bt_login = login_window.child_window(auto_id="btnLogin")
        #
        # if input_username.window_text() == "":
        #     input_username.set_focus()  # 确保焦点
        #     input_username.type_keys(config['user'])  # 输入用户名
        #
        # # 更安全的密码输入方式（避免明文记录）
        # input_pwd.set_focus()
        # input_pwd.type_keys(config['pwd'] + "{ENTER}", with_spaces=True)  # 支持特殊键

    ################## 进入主界面 ##################
    # app = Application(backend="win32").connect(title="分销系统", timeout=10)
    #
    # main_window = app.window(title="分销系统")
    # main_window.wait("visible", timeout=30)
    #
    # util.system.print_inline("正在初始化应用程序窗口...", color="green")
    # main_window.set_focus()
    # main_window.move_window(x=0, y=0, width=1110, height=700, repaint=True)
    # time.sleep(1)

    # if config['currentPage'] <= 0:
        # ################## 操作主界面，进入指定界面 ##################
        # util.system.print_inline("正在定位到指定页面...", color="blue")
        # util.system.print_inline("点击【产品】选项卡")
        # util.system.safe_click(x=296, y=40)
        # time.sleep(1)
        # util.system.print_inline("点击【亚马逊选品】选项卡")
        # util.system.safe_click(x=56, y=678)
        # time.sleep(1)
        # util.system.print_inline("点击【搜索词排名】选项卡")
        # util.system.safe_click(x=71, y=618)
        # util.app.check_load_finished()
        # util.system.print_inline("点击【日本】选项卡")
        # util.system.safe_click(x=329, y=77)
        # util.app.check_load_finished()
        # util.system.print_inline("点击【50000+】筛选条件")
        # util.system.safe_click(x=747, y=159)
        # util.app.check_load_finished()
        # util.system.print_inline("点击【排名上升+】筛选条件")
        # util.system.safe_click(x=494, y=265)
        # util.app.check_load_finished()
        # util.system.print_inline("点击【1万+】筛选条件")
        # util.system.safe_click(x=655, y=299)
        # util.app.check_load_finished()

    # total_item = 0
    # total_page = 0
    # match = re.search(r'共有\s*(\d+)\s*条记录', main_window.child_window(auto_id="lblTotal").window_text())
    # if match:
    #     total_item = int(match.group(1))
    #     total_page = math.ceil(total_item / 60)
    # if total_item <= 0:
    #     util.system.print_inline("搜索词分页数据获取失败！程序已终止运行。", color="red")
    #     util.app.press_any_key_exit()
    #     sys.exit()
    # else:
    #     util.system.print_inline(f"共获取到 {total_item} 条搜索词数据，一共 {total_page} 页", color="green")

    ################## 获取数据 ##################
    # try:
    #     util.system.print_inline("正在获取最新 Cookie 数据...")
    #
    #     resp = requests.get(
    #         "https://www.amazon.co.jp/s?k=cat",
    #         headers={
    #             "Content-Type": "text/html;charset=UTF-8",
    #             "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0"
    #         },
    #         timeout=60,
    #         stream=True # 使用 resp.cookies.get 必须开启 stream
    #     )
    #     resp.raise_for_status()
    #
    #     resp2 = requests.post(
    #         "https://www.amazon.co.jp/portal-migration/hz/glow/address-change?actionSource=glow",
    #         json={"locationType": "LOCATION_INPUT", "zipCode": "169-0074", "deviceType": "web",
    #               "storeContext": "hpc",
    #               "pageType": "Detail", "actionSource": "glow"},
    #         headers={
    #             "Content-Type": "application/json",
    #             "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0",
    #             "Cookie": resp.headers.get("set-cookie"),
    #         },
    #         timeout=60,
    #         stream=True
    #     )
    #     resp2.raise_for_status()
    #
    #     session_id = resp.cookies.get("session-id", "")
    #     ubid = resp2.cookies.get("ubid-acbjp", "")
    #     if len(session_id) == 0 or len(ubid) == 0:
    #         raise Exception("Response is ok, but values are empty")
    #     else:
    #         cookies_raw = "ubid-acbjp=" + ubid + "; session-id=" + session_id
    # except Exception as e:
    #     util.system.print_inline(f"{e}", color="red")
    #     util.system.print_inline("最新 Cookie 数据获取失败，将使用临时数据", color="yellow")
    #
    #     cookies_raw = "ubid-acbjp=355-5685452-2837352; session-id=357-7564356-4927846"

    # page_rect = main_window.child_window(auto_id="PTurn").rectangle()

    # for cur_page in range(0 if config['currentPage'] <= 0 else (config['currentPage'] - 1), total_page):
        # util.system.print_inline(f"正在获取第 {cur_page + 1}/{total_page} 页搜索词列表...")
        # main_window.set_focus()
        # time.sleep(1)
        # util.system.safe_click(x=541, y=647)
        # time.sleep(1)
        #
        # if cur_page == 0:
        #     util.system.safe_click(x=155, y=600)
        # else:
        #     while True:
        #         point = util.cv.get_next_page_point(page_rect, debug=DEBUG)
        #         if point is not None:
        #             x, y = point
        #             util.system.safe_click(x=x, y=y)
        #             break
        #         else:
        #             time.sleep(1)
        #
        # util.app.check_load_finished()
        # pyperclip.copy("")
        # util.system.safe_right_click(x=233, y=464)
        # time.sleep(1)
        # util.system.safe_click(x=287, y=504)
        # time.sleep(1)
        # kw_list = pyperclip.paste().splitlines()
        # util.system.print_inline(f"成功获取到 {len(kw_list)} 个搜索词", color="blue")

        # saved_kw = []
        # i = 0
        # for kw in kw_list:
        #     i += 1
        #     util.system.print_inline(f"当前正在筛选搜索词 (第 {cur_page + 1}/{total_page} 页, 第 {i}/{len(kw_list)}) 个：" + kw,
        #                  color="yellow")  # , newline=False

            # try:
            #     response = requests.get(
            #         "https://www.amazon.co.jp/s?k=" + kw.strip() + "&language=zh_CN",
            #         timeout=60,
            #         headers={
            #             "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0",
            #             "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            #             "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
            #             "Accept-Encoding": "gzip, deflate, br, zstd",
            #             "Cookie": cookies_raw,
            #         },
            #     )
            #     response.raise_for_status()
            #
            #     res_html = response.text
            # except Exception as e:
            #     util.system.print_inline(str(e), color="red")
            #     util.system.print_inline("获取失败，自动筛选下一个搜索词...（若频繁超时，请检查本地网络与VPN）", color="red")
            #     time.sleep(config["fetchDelay"])
            #     continue
            #
            # soup = BeautifulSoup(res_html, "html.parser")

            # try:
            #     if "ご迷惑をおかけしています" in soup.title.string:
            #         util.system.print_inline("遇到亚马逊风控，访问终止，自动筛选下一个搜索词...", color="red")
            #         time.sleep(config["fetchDelay"])
            #         continue
            #
            #     if "Amazon.co.jp" not in soup.title.string:
            #         util.system.print_inline("注意，获取的网页结果可能存在问题", color="yellow")
            #
            #     els = soup.select("div.udm-primary-delivery-message span.a-text-bold")
            # except Exception as e:
            #     util.system.print_inline("抓取该搜索词时发生错误：" + str(e), color="red")
            #     els = []

            # count = 0
            # for el in els:
            #     match = re.search(r'(\d{1,2})月(\d{1,2})日', el.get_text())
            #     if match:
            #         month = int(match.group(1))
            #         day = int(match.group(2))
            #
            #         today = date.today()
            #         current_year = today.year
            #
            #         try:
            #             target_date = date(current_year, month, day)
            #
            #             delta = (target_date - today).days
            #             if delta >= config['minDateInterval']:
            #                 count += 1
            #         except ValueError:
            #             continue
            #             # util.system.print_inline("非法日期，比如 2月30日 之类", color="red")
            #     else:
            #         continue
            #         # util.system.print_inline("未匹配到日期", color="red")

            # if count >= config['matchCount']:
            #     util.system.print_inline("该搜索词符合预期条件，已记录", color="green")  # , newline=False
            #
            #     img = None
            #     price_symbol = None
            #     price = None
            #     buy_number = None
            #     if len(els) > 0:
            #         util.system.print_inline("正在获取该搜索词对应的详细信息...", color="yellow")
            #
            #         elements = soup.select('div[cel_widget_id*="MAIN-SEARCH_RESULTS-"]')
            #
            #         for element in elements:
            #             if img is not None and price_symbol is not None and price is not None and buy_number is not None:
            #                 break
            #
            #             if element.select('div.udm-primary-delivery-message span.a-text-bold') is None:
            #                 continue
            #
            #             if img is None:
            #                 img_el = element.find('img', class_='s-image')
            #                 img_result = img_el['src'] if img_el and 'src' in img_el.attrs else None
            #                 if img_result is not None and img_result.strip().startswith('http'):
            #                     img = img_result
            #                     util.system.print_inline("图片获取成功：" + img, color="green")
            #
            #             if price_symbol is None or price is None:
            #                 price_el = element.select_one('a[aria-describedby="price-link"]')
            #                 if price_el:
            #                     if price_symbol is None:
            #                         p_symbol_el = price_el.find('span', class_='a-price-symbol')
            #                         p_symbol = p_symbol_el.get_text().strip() if p_symbol_el else ''
            #                         if len(p_symbol) > 0:
            #                             price_symbol = p_symbol
            #                             util.system.print_inline("货币类型获取成功：" + price_symbol, color="green")
            #
            #                     if price is None:
            #                         p_whole_el = price_el.find('span', class_='a-price-whole')
            #                         if p_whole_el:
            #                             p_whole = p_whole_el.get_text().strip()
            #                             p_decimal_el = p_whole_el.find('span', class_='a-price-decimal')
            #                             p_decimal = p_decimal_el.text.strip() if p_decimal_el else ''
            #                             try:
            #                                 price = int(p_whole.replace(',', '').replace(p_decimal, ''))
            #                                 util.system.print_inline("价格获取成功：" + str(price), color="green")
            #                             except Exception:
            #                                 pass
            #
            #             if buy_number is None:
            #                 buy_number_el = element.select_one('div[data-cy="reviews-block"]')
            #                 if buy_number_el:
            #                     b_number_el = buy_number_el.select_one('span.a-size-base.a-color-secondary')
            #                     b_number_full = b_number_el.get_text().strip() if b_number_el else ''
            #
            #                     # 匹配模式：
            #                     # (\d+)  —— 数字部分（整数）
            #                     # \s*              —— 允许数字和单位之间有空格
            #                     # (万|百万|千万|亿)? —— 单位，可选
            #                     match = re.search(r'过去一个月有(\d+)\s*(万|百万|千万|亿)?\+?', b_number_full)
            #                     if match:
            #                         try:
            #                             num_str, unit = match.groups()
            #                             num = int(num_str)
            #
            #                             # 单位换算表
            #                             unit_map = {
            #                                 None: 1,
            #                                 '万': 10_000,
            #                                 '百万': 1_000_000,
            #                                 '千万': 10_000_000,
            #                                 '亿': 100_000_000
            #                             }
            #
            #                             factor = unit_map.get(unit, 1)
            #                             buy_number = int(num * factor)
            #                             util.system.print_inline("购买数量获取成功：" + str(buy_number), color="green")
            #                         except Exception:
            #                             pass
            #         else:
            #             util.system.print_inline("图片或价格等信息获取失败，自动跳过（不影响流程，可忽略）", color="red")
            #
            #     saved_kw.append((kw.strip(), img, price_symbol, price, buy_number))
            # else:
            #     util.system.print_inline("该搜索词不符合预期条件，已忽略", color="red")  # , newline=False
            # time.sleep(config["fetchDelay"])

        # 将该页的数据保存为 Excel
        # if len(saved_kw) <= 0:
        #     util.system.print_inline("当前页筛选后的搜索词数量为 0，跳过保存为 Excel 的步骤", color="yellow")
        #     continue
        # else:
        #     util.system.print_inline(f"当前页筛选后的搜索词数量为 {len(saved_kw)} 个，正在生成 Excel 文件...", color="yellow")
        #
        # try:
        #     # 创建工作簿和工作表
        #     wb = Workbook()
        #     ws = wb.active
        #     ws.title = "亚马逊搜索词"
        #
        #     # 设置标题
        #     title = "亚马逊 [日本]区域 搜索词"
        #     ws['A1'] = title
        #     ws['A1'].font = Font(bold=True)
        #
        #     # 列标题（第二行）
        #     ws.cell(row=2, column=1, value="搜索词").font = Font(bold=True)
        #     ws.cell(row=2, column=2, value="图片").font = Font(bold=True)
        #
        #     # 写入数据和超链接 & 图片
        #     for i, (text, img_url, _, _, _) in enumerate(saved_kw, start=3):  # 从第3行开始写
        #         # 第一列：搜索词 + 超链接
        #         cell = ws.cell(row=i, column=1, value=text)
        #         cell.hyperlink = "https://www.amazon.co.jp/s?k=" + text
        #         cell.style = "Hyperlink"
        #
        #         # 第二列：下载图片并插入
        #         if img_url is None:
        #             continue
        #
        #         try:
        #             resp = requests.get(img_url, timeout=10)
        #             if resp.status_code == 200:
        #                 img_data = BytesIO(resp.content)
        #                 img = XLImage(img_data)
        #                 img.width, img.height = 80, 80  # 缩放图片大小
        #                 img_cell = f"B{i}"
        #                 ws.add_image(img, img_cell)
        #                 ws.row_dimensions[i].height = 65  # 调整行高
        #         except Exception as e:
        #             util.system.print_inline(f"搜索词 {text} 的图片下载失败，已忽略: {img_url}\n{e}", color="red")
        #
        #     # 自动调整第一列列宽
        #     max_len = max(len(text) for (text, _, _, _, _) in saved_kw)
        #     ws.column_dimensions[get_column_letter(1)].width = max_len + 5
        #     ws.column_dimensions[get_column_letter(2)].width = 15  # 第二列图片列
        #
        #     # 保存文件，文件名使用当前日期
        #     filename = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") + f"-第{cur_page + 1}页" + ".xlsx"
        #     output_dir = Path(
        #         config['excelPath'] if config['excelPath'] != "" else os.path.join(util.system.get_exe_dir(), "excel"))
        #     output_dir.mkdir(parents=True, exist_ok=True)
        #     wb.save(output_dir / filename)
        #     util.system.print_inline(f"已将数据存储为 Excel 文件：{filename}", color="green")
        # except Exception as e:
        #     util.system.print_inline(str(e), color="red")
        #     util.system.print_inline("Excel 生成失败！", color="red")

    #     # 数据上传到服务器
    #     ex = util.net.save_kw_to_server(saved_kw)
    #     if ex is None:
    #         util.system.print_inline("** 数据已同步上传到服务器端 **", color="green")
    #     else:
    #         util.system.print_inline(f"** 数据上传服务器失败：{ex}（不影响流程，可忽略） **", color="yellow")
    #
    # util.system.print_inline("程序执行完毕！", color="green")
    # util.app.press_any_key_exit()


# if __name__ == '__main__':
#     util.app.ensure_start_by_self()
#
#     try:
#         main()
#     except Exception as e:
#         print(e)
#         util.system.print_inline("程序执行出错！", color="red")
#         util.app.press_any_key_exit()
